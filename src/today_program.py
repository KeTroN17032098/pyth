import json
from tkinter import *
import time
import os.path
import tkinter.ttk as ttk
from tkinter import messagebox
from tkinter.filedialog import *
from tkinter import scrolledtext
import tkinter.font as tkFont
from tkcalendar import *
import tkinter as tk
import datetime
from datetime import timedelta
from PIL import Image, ImageTk
import getpass as gp
from multiprocessing import Process, Queue
import threading
import pprint
import pandas as pd
import openpyxl
from openpyxl.styles import Font,Alignment,PatternFill,Color
import shutil
from openpyxl.styles.borders import Border, Side
import webbrowser
import sys
import winshell
import schedule



today_file_path='data/today_result.json'
history_file_path='data/today_history.json'
setting_file_path='data/setting.json'
icon_path='checkbox/logo.ico'
excel_path='data_excel.xlsx'
history_excel_path='history_excel.xlsx'
images_path=['places/notebook.png','places/print.png','places/dvd.png']

FILE_CLOSE = FALSE

def ctrlEvent(event):#crtl+c제외 키 블락
    if(12==event.state and event.keysym=='c' ):
        return
    else:
        return "break"

def date_range(start, end):
    start = datetime.datetime.strptime(start, "%Y-%m-%d")
    end = datetime.datetime.strptime(end, "%Y-%m-%d")
    dates = [date.strftime("%Y-%m-%d") for date in pd.date_range(start, periods=(end-start).days+1)]
    return dates

def today_date_str():
    return datetime.datetime.today().strftime("%Y-%m-%d")

def now_time_str():
    return today_date_str()+datetime.datetime.now().strftime(" %I:%M:%S %p")

def imageModifier(FILE_PATH,x,y):#아미지 사이즈 재설정 함수
    img1=Image.open(FILE_PATH)
    img1=img1.resize((x,y),Image.ANTIALIAS)
    resized_img1=ImageTk.PhotoImage(img1)
    return resized_img1

    
class Json_Data():
    def __init__(
        self,FilePath=today_file_path,
        HistoryPath=history_file_path,
        SettingPath=setting_file_path,
        iconpath=icon_path,
        whereList=["노트북",'프린트','관내열람'],
        genderList=['남성','여성'],
        imagelist=images_path,
        savetime='17:50',
        savedir=winshell.desktop()):#생성자
        
        if os.path.isfile(iconpath):pass#아이콘 체크
        else:iconpath=''
        
        for image in imagelist:#이미지 리스트 체크
            if os.path.isfile(image):pass
            else:imagelist.remove(image)
        
        self.SettingPath=SettingPath
        self.settings={}
        
        self.load_settings()#세팅 파일 로드
        self.check_setting(FilePath,HistoryPath,whereList,genderList,imagelist,savetime,savedir,iconpath)
        

        self.__Today_Data__=[]
        self.data={}
        self.history={}

        if self.check_save():
            with open(self.FilePath,"r") as f:
                self.data = json.load(f)
            self.check_date()
            if self.check_his():
                with open(self.HistoryPath,"r") as h:
                    self.history=json.load(h)
                    self.add_history_member(type='init')
            else:
                self.history['init']=[]
                self.history['push']=[]
                self.history['show']=[]
                self.history['built']=[]
                self.history['fail']=[]
                self.add_history_member(type='built')
                self.add_history_member(type='init')
        else:
            if self.check_his():
                with open(self.HistoryPath,"r") as h:
                    self.history=json.load(h)
                    self.add_history_member(type='built')
                    self.add_history_member(type='init')
            else:
                self.history['init']=[]
                self.history['push']=[]
                self.history['show']=[]
                self.history['built']=[]
                self.history['fail']=[]
                self.add_history_member(type='built')
                self.add_history_member(type='init')
            self.data['cumulative']=0#전체 누적 이용자 수
            self.data['typecode']="TR"#데이터 타입
            for where in self.__WhereName__:
                self.data[where]=[]
                k={
                    "date":today_date_str(),
                    "where":where
                }
                for gender in self.__Gender__:
                    k[gender]=0
                self.data[where].append(k)
            self.check_date()
            self.save_data()
            
    def save_data(self):#데이터 저장
        with open(self.FilePath,'w') as fk:
            json.dump(self.data, fk,indent=4)
    
    def load_settings(self):
        if os.path.isfile(self.SettingPath):
            with open(self.SettingPath,'r') as stpr:
                self.settings=json.load(stpr)
            
    def check_setting(self,FilePath,HistoryPath,whereList,genderList,imagelist,savetime,savedir,iconpath):
        '''
        세팅 파일에서 읽어온 데이터 체크
        ->만약 내용이 없다면 1번(새로 만들고 저장 후 자기함수호출)
        ->내용이 불완전하다 2번(세팅 읽어온거 다 지우고 자기함수호출)
        ->내용이 맞다면 3번(세이브 파일 요소로 등록 후 함수 나옴)
        '''
        if self.settings=={}:
            self.settings['FilePath']=[]
            self.settings['FilePath'].append({
                'path':FilePath,
                'type':'Data'
            })
            self.settings['FilePath'].append({
                'path':HistoryPath,
                'type':'History'
            })
            self.settings['NameList']=[]
            self.settings['NameList'].append({
                'list':whereList,
                'type':'Place'
            })
            self.settings['NameList'].append({
                'list':genderList,
                'type':'Gender'
            })
            self.settings['ImagePath']=imagelist
            self.settings['DailySave']={
                'time':savetime,
                'dir':os.path.abspath(savedir)
            }
            self.settings['IconPath']=iconpath
            self.save_settings()
            self.check_setting(FilePath,HistoryPath,whereList,genderList,imagelist,savetime,savedir,iconpath)
        else:
            if sorted(list(self.settings.keys()))!=sorted(['FilePath','NameList','ImagePath','DailySave','IconPath']):
                self.settings={}
                self.check_setting(FilePath,HistoryPath,whereList,genderList,imagelist,savetime,savedir,iconpath)
            else:
                self.settings_to_self()
    
    def settings_to_self(self):#세팅 dict로부터 요소들 읽어오기
        for d in self.settings['FilePath']:
            if d['type']=='Data':self.FilePath=d['path']
            elif d['type']=='History':self.HistoryPath=d['path']
        for f in self.settings['NameList']:
            if f['type']=='Place': self.__WhereName__=f['list']
            elif f['type']=='Gender':self.__Gender__=f['list']
        self.imagelist=self.settings['ImagePath']
        self.savetime=self.settings['DailySave']['time']
        self.savedir=self.settings['DailySave']['dir']
        self.icon_path=self.settings['IconPath']
     
    def change_settings(self,**kwargs):#요소를 바꾸고 세팅 dict에 넣은 후 저장, 바뀐 것이 있가면 True 리턴
        ischanged=False
        for k,v in kwargs.items():
            print (k)
            print(v)
            print('===========================')
            k=k.lower()
            if k=='savedir' and os.path.isdir(v):
                print('change savedir')
                self.savedir=v
                self.settings['DailySave']['dir']=self.savedir
                self.save_settings()
                ischanged=True
            if k=='iconpath' and os.path.isfile(v):
                print('change icon path')
                self.icon_path=v
                self.settings['IconPath']=self.icon_path
                self.save_settings()
                ischanged=True
            if k=='savetime' and v is str:
                try:
                    time.strftime(v,'%H:%M')
                    print('change savetime')
                    self.savetime=v
                    self.settings['DailySave']['time']=self.savetime
                    self.save_settings()
                    ischanged=True
                except ValueError:
                    pass
            if k=='imagelist' and type(v)==list:
                for vd in v:#이미지 리스트 체크
                    if os.path.isfile(vd):pass
                    else:v.remove(vd)
                print('change imagelist')
                self.imagelist=v
                self.settings['ImagePath']=self.imagelist
                self.save_settings()
                ischanged=True
            if k=='wherelist' and type(v)==list:
                for vd in v:
                    if type(vd) is not str:
                        v.remove(vd)
                print('change places name')
                self.__WhereName__=v
                for f in self.settings['NameList']:
                    if f['type']=='Place': f['list']=self.__WhereName__
                self.save_settings()
                ischanged=True
            if k=='genderlist' and type(v)==list:
                for vd in v:
                    if type(vd) is not str:
                        v.remove(vd)
                print('change Gender name')
                self.__Gender__=v
                for f in self.settings['NameList']:
                    if f['type']=='Gender': f['list']=self.__Gender__
                self.save_settings()
                ischanged=True
            if k=='filepath':
                print('Change FilePath')
                self.FilePath=v
                for d in self.settings['FilePath']:
                    if d['type']=='Data':d['path']=self.FilePath
                self.save_settings()
                ischanged=True
            if k=='historypath':
                print('Change HistoryPath')
                self.HistoryPath=v
                for d in self.settings['FilePath']:
                    if d['type']=='History':d['path']=self.HistoryPath
                self.save_settings()
                ischanged=True
        return ischanged 
    
    def show_settings(self,*args):
        tmp={}
        for arg in args:
            k=None
            if arg.lower()=='filepath':k=self.FilePath
            elif arg.lower()=='historypath':k=self.HistoryPath
            elif arg.lower()=='wherelist':k=self.__WhereName__
            elif arg.lower()=='genderlist':k=self.__Gender__
            elif arg.lower()=='imagelist':k=self.imagelist
            elif arg.lower()=='savetime':k=self.savetime
            elif arg.lower()=='savedir':k=self.savedir
            elif arg.lower()=='iconpath':k=self.icon_path
            if k!=None:tmp[arg]=k
        return tmp
    def save_settings(self):
        with open(self.SettingPath,'w') as kosk:
            json.dump(self.settings,kosk, indent=4)
    
    def save_history(self):#내역 저장
        with open(self.HistoryPath,'w') as hk:
            json.dump(self.history,hk,indent=4)

    def add_history_member(self,type='push',count=1,gender="",place="",showdates=(today_date_str(),today_date_str())):
        tmp={}
        if type =='built'or type=='init':
            tmp['user']=gp.getuser()
            tmp['time']=now_time_str()
            tmp['type']=type
            tmp['isSucess']=TRUE
        elif type=='push' and gender in self.__Gender__ and place in self.__WhereName__:
            tmp['user']=gp.getuser()
            tmp['time']=now_time_str()
            tmp['type']=type
            tmp['isSucess']=TRUE
            tmp['gender']=gender
            tmp['place']=place
            tmp['count']=count
        elif type=='show' and len(date_range(showdates[0],showdates[1]))>0:
            tmp['user']=gp.getuser()
            tmp['time']=now_time_str()
            tmp['type']=type
            tmp['isSucess']=TRUE
            tmp['dates']=date_range(showdates[0],showdates[1])
        else:
            tmp['user']=gp.getuser()
            tmp['time']=now_time_str()
            tmp['type']=type
            tmp['isSucess']=FALSE
            tmp['count']=count
            if gender!="":tmp['gender']=gender
            if place!="":tmp['place']=place
            self.history['fail'].append(tmp)
            self.save_history()
            return FALSE
        self.history[tmp['type']].append(tmp)
        self.save_history()
        return TRUE

    def check_date(self,date_w=today_date_str()):#오늘 날짜 멤버 체크 및 생성
        today=date_w
        print('check_date')
        print(today)
        self.__Today_Data__=[]
        for where in self.__WhereName__:
            for item in self.data[where]:
                if today==item['date']: 
                    self.__Today_Data__.append(item)
                    break
        print('sada')
        print(self.__Today_Data__)
        if len(self.__Today_Data__)<len(self.__WhereName__):
            hasMember=[]
            for item in self.__Today_Data__:
                hasMember.append(item['where'])
            tmp=list(set(self.__WhereName__)-set(hasMember))
            print('fdasfasd')
            print(tmp)
            for place in tmp:
                kal={
                    "date":today,
                    "where":place
                }
                for gender in self.__Gender__:
                    kal[gender]=0
                self.data[place].append(kal)
            self.check_date(date_w=today)
        elif len(self.__WhereName__)==len(self.__Today_Data__):
            self.save_data()
            return
        else: self.check_date(date_w=today)

    def show_data(self,start=today_date_str(),end=today_date_str(),mode=FALSE):#특정 기간의 데이터 검색
        end_date=datetime.datetime.strptime(end, '%Y-%m-%d')
        start_date=datetime.datetime.strptime(start, '%Y-%m-%d')
        result=[]
        delta=end_date - start_date
        if delta.days==0:
            for place in self.__WhereName__:
                for data in self.data[place]:
                    if datetime.datetime.strptime(data['date'],'%Y-%m-%d')==start_date:
                        result.append(data)
                        break
        elif delta.days>=1:
            for time in [start_date+timedelta(days=i) for i in range(delta.days+1)]:
                for place in self.__WhereName__:
                    for data in self.data[place]:
                        if datetime.datetime.strptime(data['date'],'%Y-%m-%d')==time:
                            result.append(data)
                            break
        if mode:self.add_history_member(type='show',showdates=(start,end))
        return result
                
    def push_data(self,date=today_date_str(),gender="",where="",count=1):#날짜/성별/장소 를 지정 count만큼 추가
        sucess=FALSE
        print('hey')
        if gender not in self.__Gender__:pass#Invalid Gender
        elif where not in self.__WhereName__:pass#Invalid Place
        elif [data for data in self.show_data(date,date) if data['where']==where]==[]:
            print(self.show_data(date,date))#No Member0
            print(where)
        else:
            target=[data for data in self.show_data(date,date) if data['where']==where][0]
            print('kola')
            print(target)
            for member in self.data[where]:
                if member==target:
                    member[gender]+=count
                    if member[gender]<0:
                        member[gender]=0
                    self.data['cumulative']+=count
                    if self.data['cumulative']<=0:
                        self.data['cumulative']=0
                    sucess=TRUE
                    break
        self.save_data()
        self.add_history_member(type='push',gender=gender,place=where,count=count)
        return sucess

    def modify_history_excel(self,excel_path=history_excel_path):
        """
        excel_path : if mode is TRUE excel file will save at excel_path. Defalt :'history_excel.xlsx'
        """
        columns_h={}
        for key in list(self.history.keys()):
            try:
                columns_h[key]=list(self.history[key][0].keys())
            except IndexError:
                columns_h[key]=[]
                
        wb=openpyxl.Workbook()
        for key in list(self.history.keys()):
            ws=wb.create_sheet(key)
            if len(columns_h[key])!=0:
                for cl in range(len(columns_h[key])):
                    ws.cell(row=1,column=cl+2).value=columns_h[key][cl]
                    ws.cell(row=1,column=cl+2).alignment=Alignment(horizontal='center',vertical='center')
                    ws.cell(row=1,column=cl+2).fill=PatternFill(patternType='solid',fgColor=Color('FFC000'))
                for index in range(len(self.history[key])):
                    ws.cell(row=index+2,column=1).value=index+1
                    ws.cell(row=index+2,column=1).alignment=Alignment(horizontal='center',vertical='center')
                    ws.cell(row=index+2,column=1).fill=PatternFill(patternType='solid',fgColor=Color('008000'))
                    ws.cell(row=index+2,column=1).font=Font(bold=TRUE,color=Color('ffffff'))
                    for cl in range(len(columns_h[key])):
                        if type(self.history[key][index][columns_h[key][cl]])!=list:ws.cell(row=index+2,column=cl+2).value=self.history[key][index][columns_h[key][cl]]
                        else:ws.cell(row=index+2,column=cl+2).value=str(self.history[key][index][columns_h[key][cl]])
                        ws.cell(row=index+2,column=cl+2).alignment=Alignment(horizontal='center',vertical='center')
                self.adjust_cell_width(ws)
            else:
                pass
        wb.remove(wb['Sheet'])
        try:
            wb.save(excel_path)
            os.startfile(excel_path)
        except PermissionError:
            messagebox.showerror("PermissionError","파일을 닫아주세요.")

    def adjust_cell_width(self, ws,value=1.2,cl=['A'],ri=[1]):
        """
        Adjust cell width/make (.1)&(A.) thick
        ws : worksheet handle from openpyxl
        cl : options to make other columns thick default : ['a']
        ri : options to make other rows thick default : [1]
        """
        thick_border=Border(left=Side(style='thick'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))
        for col in ws.columns:
            max_length =0
            column_col=col[0].column_letter
            for cell in col:
                cell.alignment=Alignment(horizontal='center',vertical='center')
                try:
                    if cell.column_letter in cl or cell.row in ri:cell.border=thick_border
                    if cell.value==0:cell.value='-'
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    print('kr')
                    pass
            adjusted_width=(max_length+2)*value
            print(column_col)
            print(int(adjusted_width))
            ws.column_dimensions[str(column_col)].width=int(adjusted_width)
        
    def give_savedir(self):#만약 프로그램을 다른 컴퓨터에서 실행한다면 데스크톱 유저 디렉토리도 바뀌므로 그에 대비해서 만든 함수
        tmp=''
        if os.path.isdir(self.savedir):
            pass
        elif self.savedir.split('\\')[-1]=='Desktop':
            tmp=winshell.desktop()
        else:
            tmp='data'
        if tmp!='':self.change_settings(savedir=tmp)
        return self.savedir
    
    def modify_data_excelike(self,selected_dates=[],mode=TRUE,excel_path=excel_path):
        """
        mode : make Excel file default : True
        selected_dates : list of dates that required for showing info DeFault : show_everything
        excel_path : if mode is TRUE excel file will save at excel_path. Defalt :'data_excel.xlsx'
        """
        datesd=[]
        for place in self.__WhereName__:
            for data in self.data[place]:
                if data['date'] not in datesd:datesd.append(data['date'])
        dates=[]
        if selected_dates!=[]:
            for date in datesd:
                if date in selected_dates:
                    dates.append(date)
        else:dates=datesd
        DATE=sorted(dates)
        dates=DATE
        result={}
        columns_d=["Date"]
        columns_name=[]
        for where in self.__WhereName__:
            for gender in self.__Gender__:
                columns_name.append(where+" "+gender)
        columns_d+=columns_name
        columns_d+=['총합']
        for date in range(len(dates)):
            sd=self.show_data(start=dates[date],end=dates[date],mode=TRUE)
            tmp=[dates[date]]
            for place in self.__WhereName__:
                for gender in self.__Gender__:
                    tmp.append(0)
            tmp.append(0)
            for member in sd:
                for i in range(len(self.__WhereName__)):
                    if member['where']==self.__WhereName__[i]:
                        for j in range(len(self.__Gender__)):
                            tmp[i*2+j+1]+=member[self.__Gender__[j]]
                            tmp[-1]+=tmp[i*2+j+1]
            result[str(date+1)]=tmp
        pprint.pprint(result)
        if mode:
            df=pd.DataFrame.from_dict(result,orient="index",columns=columns_d)
            try:
                df.to_excel(excel_path)
                wb=openpyxl.load_workbook(excel_path,data_only=True)
                ws=wb['Sheet1']
                sum_row=[]
                for c in ws.columns:
                    column_ind=c[0].column_letter
                    c[0].alignment=Alignment(horizontal='center',vertical='center')
                    c[0].fill=PatternFill(patternType='solid',fgColor=Color('FFC000'))
                    if column_ind !='A' and column_ind !='B':
                        tmp_add=0
                        for cell in c:
                            cell.alignment=Alignment(horizontal='center',vertical='center')
                            if cell.row!=1:
                                tmp_add+=int(cell.value)
                        sum_row.append(tmp_add)
                    elif column_ind=='A':
                        F=Font(bold=TRUE,color=Color('ffffff'))
                        for cell in c:
                            cell.alignment=Alignment(horizontal='center',vertical='center')
                            cell.fill=PatternFill(patternType='solid',fgColor=Color('008000'))
                            cell.font=F
                    else:
                        F=Font(bold=TRUE,color=Color('ffffff'))
                        for cell in c:
                            cell.alignment=Alignment(horizontal='center',vertical='center')
                            cell.fill=PatternFill(patternType='solid',fgColor=Color('008000'))
                            cell.font=F
                ws.append(["총합"," "]+sum_row)
                kolase=[]
                for kas in range(len(self.__Gender__)):
                    j=0
                    for suro in range(len(sum_row)):
                        if suro%len(self.__Gender__)==kas:
                            j+=sum_row[suro]
                    kolase.append(j)
                kolase[0]-=sum_row[-1]
                kartli=[]
                for genko in range(len(self.__Gender__)):
                    kartli.append(self.__Gender__[genko]+ " 총합 :")
                    kartli.append(kolase[genko])
                place_total=[]
                GT=len(self.__Gender__)
                for plag in range(len(self.__WhereName__)):
                    loki=0
                    for geni in range(GT):
                        loki+=sum_row[plag*GT+geni]
                    place_total.append(self.__WhereName__[plag]+" 총합 :")
                    place_total.append(loki)
                ws.append(["",""]+kartli)
                ws.append(["",""]+place_total)
                self.adjust_cell_width(ws,value=2.4,cl=['A','B'])
                ws.column_dimensions['A'].width=5
                wb.save(excel_path)
                os.startfile(excel_path)
            except PermissionError:
                messagebox.showerror("PermissionError","파일을 닫아주세요.")
        return result

    def columns_name(self):
        columns_name=[]
        for where in self.__WhereName__:
            for gender in self.__Gender__:
                columns_name.append(where+" "+gender)
        return columns_name
    
    def button_event(self,event,where="",gender="",mode=TRUE):
        if where not in self.__WhereName__ or gender not in self.__Gender__:pass
        else:
            if mode==TRUE:self.push_data(gender=gender,where=where)
            else:self.push_data(gender=gender,where=where,count=-1)

    def check_save(self,path=""):
        isValid=FALSE
        if path=="":path=self.FilePath
        if os.path.isfile(path):
            check_tmp={}
            with open(path,'r') as check:
                check_tmp=json.load(check)
            if sorted(list(check_tmp.keys()))==sorted(self.__WhereName__+['cumulative','typecode']) and sorted(list(check_tmp[self.__WhereName__[0]][0].keys()))==sorted(list(self.__Gender__+['date','where'])):
                isValid=TRUE
                print("세이브 파일 무결성 확인")
        else:
            pass
        return isValid

    def check_his(self,path=""):
        isValid=FALSE
        if path=="":path=self.HistoryPath
        if os.path.isfile(path):
            check_tmp={}
            with open(path,'r') as check:
                check_tmp=json.load(check)
            if sorted(list(check_tmp.keys()))==sorted(['init','built','show','push','fail']):
                isValid=TRUE
                print("기록 파일 무결성 확인")
        else:
            pass
        return isValid

    def quit(self,mode=TRUE):
        global FILE_CLOSE
        FILE_CLOSE=mode
        print("자동저장 후 종료")
        self.save_data()
        self.save_history()
        self.save_settings()
        print('저장 완료')
        del self



class Table(ttk.Treeview):
    def __init__(self,master=None,columns=[],WIDTH=180,MINWIDTH=165,first_column="Date(날짜)"):
        super().__init__(master)
        self.master = master
        self.WIDTH=WIDTH
        self.MINWIDTH=MINWIDTH
        self.set_columns(columns=columns,first_column=first_column)
        self.set_FontStyle()
        
    def set_columns(self,columns=[],first_column="sample"):
        self.column('#0',width=self.WIDTH,minwidth=self.MINWIDTH)
        self.heading('#0',text=first_column,anchor=tk.W)
        if columns==[]:
            return
        if len(columns)==1:
            return
        else:
            self.add_columns(columns)
            for key in range(len(columns)):
                self.column(columns[key],width=self.WIDTH,minwidth=self.MINWIDTH,anchor=tk.N)
    def add_columns(self,columns,**kwargs):
        current_columns=list(self['columns'])
        current_columns={key:self.heading(key) for key in current_columns}

        self['columns']=list(current_columns.keys())+list(columns)
        for key in columns:
            self.heading(key,text=key,**kwargs)
        for key in current_columns:
            # State is not valid to set with heading
            state = current_columns[key].pop('state')
            self.heading(key, **current_columns[key]) 

    def set_FontStyle(self,fontStyle_I=("Lucida Grande",18),fontStyle_H=("Lucida Grande",18,'bold')):
            style=ttk.Style()
            style.configure("tp.Treeview",font=fontStyle_I)
            style.configure("tp.Treeview.Heading",font=fontStyle_H)
            style.configure("tp.Treeview",rowheight=25)
            self.configure(style="tp.Treeview")

    def insert_data_from_json(self,date_s=today_date_str(),date_f=today_date_str(),json_data=None,allmode=FALSE):
        if type(json_data)!=Json_Data:
            return TypeError
        elif allmode:
            tmp=json_data.modify_data_excelike(mode=FALSE)
            print('TMP')
            print(tmp)
            for data in range(len(tmp)):
                date=tmp[str(data+1)].pop(0)
                for kas in range(len(tmp[str(data+1)])):
                    if tmp[str(data+1)][kas]==0:
                        tmp[str(data+1)][kas]="-"
                self.insert('','end',text=date,values=tuple(tmp[str(data+1)]),iid=str(data+1)+'번')
        else:
            tmp=json_data.modify_data_excelike(selected_dates=date_range(date_s,date_f),mode=FALSE)
            print('TMP')
            print(tmp)
            for data in range(len(tmp)):
                date=tmp[str(data+1)].pop(0)
                for kas in range(len(tmp[str(data+1)])):
                    if tmp[str(data+1)][kas]==0:
                        tmp[str(data+1)][kas]="-"
                self.insert('','end',text=date,values=tuple(tmp[str(data+1)]),iid=str(data+1)+'번')
                
    def clear_table(self):
        for i in self.get_children():
            print(i)
            self.delete(i)

class TableSet():
    def __init__(self,master=None,columns=[],WIDTH=155,MINWIDTH=145,first_column="Date(날짜)"):
        self.Table=Table(master,columns=columns,WIDTH=WIDTH,MINWIDTH=MINWIDTH,first_column=first_column)
        self.TabelYScroll=ttk.Scrollbar(master,orient="vertical",command=self.Table.yview)
        self.TableXScroll=ttk.Scrollbar(master,orient="horizontal",command=self.Table.xview)
        self.Table.configure(yscrollcommand=self.TabelYScroll.set,xscrollcommand=self.TableXScroll.set)
        self.TabelYScroll.pack(side="right",fill='y')
        self.TableXScroll.pack(side="bottom",fill='x')
        self.Table.pack(fill="x")
 
class MyDateEntry(DateEntry):
    def __init__(self, master=None, **kw):
        DateEntry.__init__(self, master=None, **kw)
        # add black border around drop-down calendar
        self._top_cal.configure(bg='black', bd=1)
        # add label displaying today's date below
        tk.Label(self._top_cal, bg='gray90', anchor='w',
                    text='Today: %s' % datetime.datetime.today().strftime('%x')).pack(fill='x')
        
class Application(Frame):
    def __init__(self,master=None,savefile=None):
        super().__init__(master)
        self.master = master
        self.savefile = savefile
        self.image_sets = self.savefile.imagelist
        self.viewmode=FALSE
        print(self.savefile.data)
        self.window_set()
        self.pack()
        self.create_menu()
        self.create_widgets()
        master.protocol("WM_DELETE_WINDOW",self.quit_all)#X버튼을 눌러서 종료시
        self.UpdateTexts(self.savefile.savetime,self.savefile.give_savedir())
        
        
    def window_set(self):
        self.master.iconbitmap(default=icon_path)
        self.master.title("전자정보실 "+today_date_str()+"일 기록")
        self.master.geometry("1200x600")
        self.fontStyle3=tkFont.Font(family="Lucida Grande",size=18,weight='bold')

    def create_menu(self):
        self.menubar=Menu(self.master)
        self.menu1=Menu(self.menubar,tearoff=0)
        self.menu1.add_command(label="오늘만",command=self.saveasexcel_today)
        self.menu1.add_command(label='날짜 지정',command=self.saveasexcel_selected)
        self.menu1.add_command(label='모두 다',command=self.saveasexcel_all)
        self.menu1.add_separator()
        self.menu1.add_command(label="변경 내역",command=self.saveasexcel_history)
        self.menubar.add_cascade(label="엑셀화",menu=self.menu1)
        self.menu2=Menu(self.menubar,tearoff=0)
        self.menu2.add_command(label='데이터 불러오기',command=self.readjson_data)
        self.menu2.add_command(label='히스토리 불러오기',command=self.readjson_history)
        self.menu2.add_separator()
        self.menu2.add_command(label='데이터 다른 이름으로 저장',command=self.saveasjson_data)
        self.menu2.add_command(label='히스토리 다른 이름으로 저장',command=self.saveasjson_history)
        self.menubar.add_cascade(label="세이브/로드",menu=self.menu2)
        self.menu4=Menu(self.menubar,tearoff=0)
        self.menu4.add_command(label="다른 날짜 수정",command=self.edit_other_date)
        self.menu4.add_command(label="설정 수정",command=self.edit_settings)
        self.menubar.add_cascade(label='수정',menu=self.menu4)
        self.menu3=Menu(self.menubar,tearoff=0)
        self.menu3.add_command(label="설명서",command=self.show_help)
        self.menu3.add_command(label="Copyrights",command=self.show_copyright)
        self.menubar.add_cascade(label="도움",menu=self.menu3)
        self.master.config(menu=self.menubar)

    def create_widgets(self):
        self.fontStyle=tkFont.Font(family="Lucida Grande",size=15)
        self.fontStyle2=tkFont.Font(family="Lucida Grande",size=25)
        self.style=ttk.Style(self.master)
        self.style.theme_use('clam')
        self.create_DateField()
        self.create_TableField()
        self.create_ButtonField()
        self.create_ViewButtonField()
        
    def create_DateField(self):
        self.dateField=Frame(self)
        self.dateField.pack(fill="x")
        
        self.DateEntry=Entry(self.dateField,font=self.fontStyle2,width=22)
        self.DateEntry.pack()
    
    def updatetable(self,event):
            self.Table.Table.clear_table()
            self.Table.Table.insert_data_from_json(json_data=self.savefile,allmode=self.viewmode)
    
    def create_TableField(self):
        self.TableField=Frame(self,relief="solid",bd=2)
        self.TableField.pack(fill="x")

        '''self.Table=ttk.Treeview(self.TableField,columns=['1','2'],displaycolumns=['1','2'])
        self.Table.pack()'''
        '''
        self.Table=Table(self.TableField,columns=self.savefile.columns_name())
        self.TabelYScroll=ttk.Scrollbar(self.TableField,orient="vertical",command=self.Table.yview)
        self.TableXScroll=ttk.Scrollbar(self.TableField,orient="horizontal",command=self.Table.xview)
        self.Table.configure(yscrollcommand=self.TabelYScroll.set,xscrollcommand=self.TableXScroll.set)
        self.TabelYScroll.pack(side="right",fill='y')
        self.TableXScroll.pack(side="bottom",fill='x')
        self.Table.pack(fill="x")
        '''
        self.Table=TableSet(self.TableField,columns=self.savefile.columns_name())
        
        
    def create_ButtonField(self):
        self.ButtonField=Frame(self,relief='solid',bd=2)
        self.ButtonField.pack(fill='x',side='bottom')

        self.SubButtonFields=[]
        self.Buttons=[]
        for place_index in range(len(self.savefile.__WhereName__)):
            SBF=Frame(self.ButtonField,relief='solid',bd=2)
            self.SubButtonFields.append(SBF)
            if len(self.image_sets)!=0:
                SBIF=Frame(self.ButtonField,relief='solid',bd=2)
                self.SubButtonFields.append(SBIF)
                SBIF.pack(fill='both',side='left')
                img=imageModifier(self.image_sets[place_index%len(self.image_sets)],200,200)
                image=Label(SBIF,image=img,background="white")
                image.image=img
                image.pack(fill='both')
            SBF.pack(fill='both',side='left')
            for gender in self.savefile.__Gender__:
                button=Button(SBF,text=self.savefile.__WhereName__[place_index]+" "+gender,font=self.fontStyle)
                button.pack()
                button.bind("<Button-1>",lambda event, w=self.savefile.__WhereName__[place_index],g=gender,m=TRUE:self.savefile.button_event(event,where=w,gender=g,mode=m))
                button.bind("<ButtonRelease-1>",self.updatetable)
                self.Buttons.append(button)
            for gender in self.savefile.__Gender__:
                button=Button(SBF,text=self.savefile.__WhereName__[place_index]+" "+gender+" 빼기",font=self.fontStyle)
                button.bind("<Button-1>",lambda event, w=self.savefile.__WhereName__[place_index],g=gender,m=FALSE:self.savefile.button_event(event,where=w,gender=g,mode=m))
                button.bind("<ButtonRelease-1>",self.updatetable)
                button.pack()
                self.Buttons.append(button)

    def create_ViewButtonField(self):
        self.ViewButtonField=Frame(self)
        self.ViewButtonField.pack()
        self.ViewButton=Button(self.ViewButtonField,text="전체/오늘 보기 변경",command=self.changemode)
        self.ViewButton.pack()

    def updateDateEntry(self):
        try:
            self.DateEntry.delete(0,END)
            now=now_time_str().center(21," ")
            self.DateEntry.insert(END,now)
            schedule.run_pending()
        except RuntimeError:
            print("런타임 에러 :쓰레드 "+threading.current_thread().name+" 종료")
            return
        except TclError:
            print("TCL 에러 :쓰레드 "+threading.current_thread().name+" 종료")
            return
        
        threading.Timer(1,self.updateDateEntry).start()
      
    def save_today_data(self,save_path):
        if save_path=="":self.savefile.modify_data_excelike(selected_dates=[today_date_str()],mode=TRUE,excel_path=os.path.join(winshell.desktop(),today_date_str()+'_exit_time.xlsx'))  
        else:self.savefile.modify_data_excelike(selected_dates=[today_date_str()],mode=TRUE,excel_path=os.path.join(os.path.abspath(save_path),today_date_str()+'_exit_time.xlsx')) 
    def changemode(self):
        if self.viewmode:
            self.viewmode=FALSE
        else:
            self.viewmode=TRUE
        self.Table.Table.clear_table()
        self.Table.Table.insert_data_from_json(json_data=self.savefile,allmode=self.viewmode)


    def UpdateTexts(self,save_time,save_path):
        print('a')
        schedule.every().day.at(save_time).do(self.save_today_data,(save_path))
        self.updateDateEntry()
        self.Table.Table.insert_data_from_json(json_data=self.savefile,allmode=self.viewmode)

    def saveasexcel_today(self):
        sv=asksaveasfilename(title="경로 지정",filetypes=[("excel file",".xlsx")],initialdir="./data",initialfile=today_date_str()+'.xlsx',defaultextension='.xlsx')
        if sv == '' or sv is None:
            return
        self.savefile.modify_data_excelike(selected_dates=[today_date_str()],mode=TRUE,excel_path=sv)
        
    def saveasexcel_all(self):
        sv=asksaveasfilename(title="경로 지정",filetypes=[("excel file",".xlsx")],initialdir="./data",initialfile='all.xlsx',defaultextension='.xlsx')
        if sv == '' or sv is None:
            return
        self.savefile.modify_data_excelike(mode=TRUE,excel_path=sv)
        
    def saveasexcel_selected(self):
        sw=Toplevel(self.master)
        sw.geometry('300x120')
        sw.resizable(False,False)
        sw.title('날짜를 선택하세요.')
        sdf=Frame(sw)
        sdf.pack()
        sdl=Label(sdf,text="Start :")
        sdl.pack(side='left')
        sde = DateEntry(sdf,selectbackground='gray80',
                 selectforeground='black',
                 normalbackground='white',
                 normalforeground='black',
                 background='gray90',
                 foreground='black',
                 bordercolor='gray90',
                 othermonthforeground='gray50',
                 othermonthbackground='white',
                 othermonthweforeground='gray50',
                 othermonthwebackground='white',
                 weekendbackground='white',
                 weekendforeground='black',
                 headersbackground='white',
                 headersforeground='gray70')
        sde.pack(side='right')
        edf=Frame(sw)
        edf.pack()
        edl=Label(edf,text="End :")
        edl.pack(side='left')
        ede = DateEntry(edf,selectbackground='gray80',
                 selectforeground='black',
                 normalbackground='white',
                 normalforeground='black',
                 background='gray90',
                 foreground='black',
                 bordercolor='gray90',
                 othermonthforeground='gray50',
                 othermonthbackground='white',
                 othermonthweforeground='gray50',
                 othermonthwebackground='white',
                 weekendbackground='white',
                 weekendforeground='black',
                 headersbackground='white',
                 headersforeground='gray70')
        ede.pack(side='right')
        def bf():
            tmp=date_range(sde.get_date().strftime("%Y-%m-%d"),ede.get_date().strftime("%Y-%m-%d"))
            if len(tmp)<=0:
                messagebox.showerror("날짜 선택","잘못된 날짜 범위 설정입니다.")
            else:
                sv=asksaveasfilename(title="경로 지정",filetypes=[("excel file",".xlsx")],initialdir="./data",initialfile=sde.get_date().strftime("%Y%m%d")+"_to_"+ede.get_date().strftime("%Y%m%d")+'.xlsx',defaultextension='.xlsx')
                if sv == '' or sv is None:
                    return
                self.savefile.modify_data_excelike(mode=TRUE,excel_path=sv,selected_dates=tmp)
        bdf=Frame(sw)
        bdf.pack()
        bdb=Button(bdf,text="엑셀화",command=bf)
        bdb.pack()

    def saveasexcel_history(self):
        sv=asksaveasfilename(title="경로 지정",filetypes=[("excel file",".xlsx")],initialdir="./data",initialfile='변경내역.xlsx',defaultextension='.xlsx')
        if sv == '' or sv is None:
            return
        self.savefile.modify_history_excel(excel_path=sv)
        
    def saveasjson_data(self):
        sv=asksaveasfilename(title="경로 지정",filetypes=[("json file(.json)",".json")],initialdir="./data",initialfile='save_data.json',defaultextension='.json')
        if sv == '' or sv is None:
            return
        with open(sv,'w') as save:
            json.dump(self.savefile.data,save,indent=4)
            
    def saveasjson_history(self):
        sv=asksaveasfilename(title="경로 지정",filetypes=[("json file(.json)",".json")],initialdir="./data",initialfile='history_data.json',defaultextension='.json')
        if sv == '' or sv is None:
            return
        with open(sv,'w') as save:
            json.dump(self.savefile.history,save,indent=4)
            
    def readjson_history(self):
        sv=askopenfilename(title="경로 지정",filetypes=[("json file(.json)",".json")],initialdir="./data")
        if sv == '' or sv is None:
            return

        if self.savefile.check_his(path=sv):
            try:
                shutil.copy2(sv,self.savefile.HistoryPath)
                del self.savefile
                self.master.destroy()
            except shutil.SameFileError:
                messagebox.showerror("파일 오류","현재 저장 파일과 동일합니다.")
                return
            
    def readjson_data(self):
        sv=askopenfilename(title="경로 지정",filetypes=[("json file(.json)",".json")],initialdir="./data")
        if sv == '' or sv is None:
            return
        if self.savefile.check_save(path=sv):
            try:
                shutil.copy2(sv,self.savefile.FilePath)
                del self.savefile
                self.master.destroy()
            except shutil.SameFileError:
                messagebox.showerror("파일 오류","현재 저장 파일과 동일합니다.")
                return
            
    def edit_other_date(self):
        subwindow=Toplevel(self.master)
        subwindow.geometry('300x80')
        subwindow.resizable(False,False)
        subwindow.title('수정')
        slb1=Label(subwindow)
        slb1.pack()
        labels1=Label(slb1,text="날짜 :")
        dates1=DateEntry(slb1,selectbackground='gray80',
                 selectforeground='black',
                 normalbackground='white',
                 normalforeground='black',
                 background='gray90',
                 foreground='black',
                 bordercolor='gray90',
                 othermonthforeground='gray50',
                 othermonthbackground='white',
                 othermonthweforeground='gray50',
                 othermonthwebackground='white',
                 weekendbackground='white',
                 weekendforeground='black',
                 headersbackground='white',
                 headersforeground='gray70')
        labels2=Label(slb1,text="대상 :")
        vbbs2=[]
        for pls in self.savefile.__WhereName__:
            for gen in self.savefile.__Gender__:
                vbbs2.append(pls+' '+gen)
        cbbs2=ttk.Combobox(slb1,values=vbbs2)
        def edit_selected_member():
            member=cbbs2.get()
            date=dates1.get_date()
            if member =='':
                messagebox.showerror('Error','Selected member in combobox')
                subwindow.lift()
                return
            elif date>datetime.datetime.now().date():
                messagebox.showerror('Error','You cannot select a date is sooner than today')
                subwindow.lift()
                return
            prev_date=date.strftime("%Y-%m-%d")
            prev_gen=member.split(" ")[1]
            prev_plc=member.split(" ")[0]
            subwindow1=Toplevel(self.master)
            tmp=self.savefile.show_data(prev_date,prev_date)
            if tmp==[]:
                j=messagebox.askyesno('No Data in certain date.','Nothing saved in selected Date.\nMake Data in certain date?')
                if j:
                    self.savefile.check_date(date_w=prev_date)
                    tmp=self.savefile.show_data(prev_date,prev_date)
                    print(tmp)
                    subwindow1.lift()
                else:
                    subwindow1.destroy()
                    subwindow.lift()
                    return
            subwindow.destroy()
            prev_value=[cant for cant in tmp if cant['where']==prev_plc][0][prev_gen]
            subwindow1.geometry('300x400')
            subwindow1.resizable(False,False)
            subwindow1.title(date.strftime("%Y-%m-%d")+' '+member+' 수정')
            slb1=Label(subwindow1)
            slb1.pack()
            lb=Label(slb1,text=date.strftime("%Y-%m-%d")+' '+member)
            lb.pack()
            ent=Entry(slb1)
            ent.insert(END,str(prev_value))
            ent.pack()
            def edit_data():
                try:
                    print(ent.get())
                    fo=int(ent.get())
                    print(fo)
                    if fo>=0:
                        if messagebox.askokcancel('Edit Data','Change '+prev_date+' '+prev_plc+' '+prev_gen+"'s count\n"+str(prev_value)+'\ninto\n'+str(fo)):
                            self.savefile.push_data(date=prev_date,gender=prev_gen,where=prev_plc,count=fo-prev_value)
                            self.updatetable(EventType.ButtonRelease)
                            self.lift()
                            subwindow1.destroy()
                    else:
                        messagebox.showerror('Error','Input Numbers without -')
                        subwindow1.lift()
                except ValueError:
                    messagebox.showerror('Error','Input Numbers')
                    subwindow1.lift()
            btn=Button(slb1,text='수정',command=edit_data)
            btn.pack()
        btns1=Button(slb1,text="선택",command=edit_selected_member)
        labels1.grid(row=0,column=0)
        dates1.grid(row=0,column=1)
        labels2.grid(row=1,column=0)
        cbbs2.grid(row=1,column=1)
        btns1.grid(row=2,column=1)
        
    
    def show_help(self):
        webbrowser.open('help.pdf')
        
    def show_copyright(self):
        webbrowser.open('copyright.pdf')    
    
    def edit_settings(self):
        FFES={}
        sw=Toplevel(self.master)
        sw.geometry('500x380')
        sw.resizable(False,False)
        sw.title('Settings')
        lbere=Label(sw,text="Save Your Data File Before change settings",font=self.fontStyle3)
        lbere.pack()
        FFES['filepath']=self.FileFindEntry(sw,type_set='filepath')
        FFES['historypath']=self.FileFindEntry(sw,type_set='historypath')
        FFES['iconpath']=self.FileFindEntry(sw,type_set='iconpath')
        FFES['savedir']=self.FileFindEntry(sw,type_set='savedir',mode=True)
        FFES['wherelist']=self.ListNameEntry(sw,type_set='wherelist')
        FFES['genderlist']=self.ListNameEntry(sw,type_set='genderlist')
        FFES['imagelist']=self.FilesFindScrolledText(sw,type_set='imagelist',file_type=[('PNG file(.png)','.png')])
        FFES['savetime']=self.TimeSetEntry(sw,type_set='savetime')
        btnr=Button(sw,text='Reset')
        def reset(event):
            if messagebox.askokcancel('Reset Settings','Your settings and your savefile will reset itself.\nAre you sure?'):
                self.savefile.settings={}
                self.savefile.quit(mode=False)
                self.master.destroy()
            else:return
        btnr.bind('<ButtonRelease-1>',reset)
        btnr.pack()
        btnk=Button(sw,text='Save')
        def change_set(event):
            if messagebox.askokcancel('Change Settings','Your savefile will reset itself.\nYour Setting will be changed.\nAre you sure?'):
                keys=FFES.keys()
                cans={}
                for key in keys:
                    if key!='imagelist':
                        lot=FFES[key]['entry'].get()
                        if key=='wherelist' or key =='genderlist':cans[key]=lot.split(',')
                        else:cans[key]=lot
                    else:cans[key]=FFES[key]['entry'].get('1.0',END).splitlines()
                self.savefile.change_settings(**cans)
                self.savefile.quit(mode=False)
                self.master.destroy()
        btnk.bind('<ButtonRelease-1>',change_set)
        btnk.pack()
        
    def FileFindEntry(self,master=None,start_dir=winshell.desktop(),file_type=[("json file(.json)",".json")],type_set='',mode=False):
        def finder(event):
            k=askopenfilename(title="경로 지정",filetypes=file_type,initialdir=start_dir)
            if k==None or k=="":return
            entry.delete(0,END)
            entry.insert(0,k)
            master.lift()
        def finder_d(event):
            k=askdirectory(title="경로 지정",initialdir=start_dir)
            if k==None or k=="":return
            entry.delete(0,END)
            entry.insert(0,k)
            master.lift()
        FFE={}  
        if master!=None:
            frame=Frame(master)
            frame.pack()
            FFE['frame']=frame
            label=Label(frame,text=type_set.upper()+' : ')
            label.grid(row=0,column=0)
            FFE['label']=label
            entry=Entry(frame)
            entry.insert(0,self.savefile.show_settings(type_set)[type_set])
            entry.bind("<Key>",lambda e: ctrlEvent(e))
            FFE['entry']=entry
            entry.grid(row=0,column=1)
            btn=Button(frame,text='Change')
            if mode:btn.bind('<ButtonRelease-1>',finder_d)
            else:btn.bind('<ButtonRelease-1>',finder)
            btn.grid(row=0,column=2)
            FFE['button']=btn
            FFE['type']=type_set
        return FFE
    
    def ListNameEntry(self,master=None,type_set=''):
        FFE={}
        if master!=None:
            frame=Frame(master)
            frame.pack()
            FFE['frame']=frame
            label=Label(frame,text=type_set.upper()+' : ')
            label.grid(row=0,column=0)
            FFE['label']=label
            entry=Entry(frame)
            tmps=''
            tmp=self.savefile.show_settings(type_set)[type_set]
            for ak in tmp:
               tmps+=','
               tmps+=ak
            tmps=tmps[1:]
            entry.insert(0,tmps) 
            FFE['entry']=entry
            entry.grid(row=0,column=1)
            FFE['type']=type_set
        return FFE
    
    def FilesFindScrolledText(self,master=None,start_dir=winshell.desktop(),file_type=[("json file(.json)",".json")],type_set=''):
        def finder(event):
            k=askopenfilenames(title="경로 지정",filetypes=file_type,initialdir=start_dir)
            if k==None or k==[""]:return
            entry.delete('1.0',END)
            for pt in k:
                entry.insert(END,pt+'\n')
            master.lift()
        FFE={}  
        if master!=None:
            frame=Frame(master)
            frame.pack()
            FFE['frame']=frame
            label=Label(frame,text=type_set.upper()+' : ')
            label.grid(row=0,column=0)
            FFE['label']=label
            entry=scrolledtext.ScrolledText(frame,height=7,width=25)
            tmp=self.savefile.show_settings(type_set)[type_set]
            for aks in tmp:
                entry.insert(END,aks+'\n')
            entry.bind("<Key>",lambda e: ctrlEvent(e))
            FFE['entry']=entry
            entry.grid(row=0,column=1)
            btn=Button(frame,text='Change')
            btn.bind('<ButtonRelease-1>',finder)
            btn.grid(row=0,column=2)
            FFE['button']=btn
            FFE['type']=type_set
        return FFE
    
    def TimeSetEntry(self,master=None,type_set=''):
        FFE={}
        if master!=None:
            frame=Frame(master)
            frame.pack()
            FFE['frame']=frame
            label=Label(frame,text=type_set.upper()+' : ')
            label.grid(row=0,column=0)
            FFE['label']=label
            entry=Entry(frame)
            tmp=self.savefile.show_settings(type_set)[type_set]
            entry.insert(0,tmp) 
            FFE['entry']=entry
            entry.grid(row=0,column=1)
            FFE['type']=type_set
        return FFE
    
    def quit_all(self):
        print("Quit All")
        self.savefile.quit()
        self.master.quit()
        self.master.destroy()

if __name__ == '__main__':#treeview 이용 오늘 뿐만 아니라 옛날 기록도 조회
    while True:
        root=Tk()
        tr=Json_Data(savetime='17:50',savedir='data')
        app=Application(master=root,savefile=tr)
        app.mainloop()
        time.sleep(5)
        if FILE_CLOSE:
            del app
            del tr
            del root
            break
    
    sys.exit()