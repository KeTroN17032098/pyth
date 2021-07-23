import PyQt5
from PyQt5 import QtCore
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
import sys
import operator
import json
from tkinter import *
import time
import os.path
import tkinter.ttk as ttk
from tkinter import messagebox
from tkinter.filedialog import *
from tkinter import scrolledtext
import tkinter.font as tkFont
from tkcalendar import *
import tkinter as tk
import datetime
from datetime import timedelta
from PIL import Image, ImageTk
import getpass as gp
from multiprocessing import Process, Queue
import threading
import pprint
import pandas as pd
import openpyxl
from openpyxl.styles import Font,Alignment,PatternFill,Color, alignment
import shutil
from openpyxl.styles.borders import Border, Side
import webbrowser
import winshell
import schedule
import pandas



today_file_path='data/today_result.json'
history_file_path='data/today_history.json'
setting_file_path='data/setting.json'
icon_path='checkbox/logo.ico'
excel_path='data_excel.xlsx'
history_excel_path='history_excel.xlsx'
images_path=['places/notebook.png','places/print.png','places/dvd.png']

FILE_CLOSE = True
    
def ctrlEvent(event):#crtl+c제외 키 블락
    if(12==event.state and event.keysym=='c' ):
        return
    else:
        return "break"

def date_range(start, end):
    start = datetime.datetime.strptime(start, "%Y-%m-%d")
    end = datetime.datetime.strptime(end, "%Y-%m-%d")
    dates = [date.strftime("%Y-%m-%d") for date in pd.date_range(start, periods=(end-start).days+1)]
    return dates

def today_date_str():
    return datetime.datetime.today().strftime("%Y-%m-%d")

def now_time_str():
    return today_date_str()+datetime.datetime.now().strftime(" %I:%M:%S %p")

def imageModifier(FILE_PATH,x,y):#아미지 사이즈 재설정 함수
    img1=Image.open(FILE_PATH)
    img1=img1.resize((x,y),Image.ANTIALIAS)
    resized_img1=ImageTk.PhotoImage(img1)
    return resized_img1

    
class Json_Data():
    def __init__(
        self,FilePath=today_file_path,
        HistoryPath=history_file_path,
        SettingPath=setting_file_path,
        iconpath=icon_path,
        whereList=["노트북",'프린트','관내열람'],
        genderList=['남성','여성'],
        imagelist=images_path,
        savetime='17:50',
        savedir=winshell.desktop()):#생성자
        
        if os.path.isfile(iconpath):pass#아이콘 체크
        else:iconpath=''
        
        for image in imagelist:#이미지 리스트 체크
            if os.path.isfile(image):pass
            else:imagelist.remove(image)
        
        self.SettingPath=SettingPath
        self.settings={}
        
        self.load_settings()#세팅 파일 로드
        self.check_setting(FilePath,HistoryPath,whereList,genderList,imagelist,savetime,savedir,iconpath)
        

        self.__Today_Data__=[]
        self.data={}
        self.history={}

        if self.check_save():
            with open(self.FilePath,"r") as f:
                self.data = json.load(f)
            self.check_date()
            if self.check_his():
                with open(self.HistoryPath,"r") as h:
                    self.history=json.load(h)
                    self.add_history_member(type='init')
            else:
                self.history['init']=[]
                self.history['push']=[]
                self.history['show']=[]
                self.history['built']=[]
                self.history['fail']=[]
                self.add_history_member(type='built')
                self.add_history_member(type='init')
        else:
            if self.check_his():
                with open(self.HistoryPath,"r") as h:
                    self.history=json.load(h)
                    self.add_history_member(type='built')
                    self.add_history_member(type='init')
            else:
                self.history['init']=[]
                self.history['push']=[]
                self.history['show']=[]
                self.history['built']=[]
                self.history['fail']=[]
                self.add_history_member(type='built')
                self.add_history_member(type='init')
            self.data['cumulative']=0#전체 누적 이용자 수
            self.data['typecode']="TR"#데이터 타입
            for where in self.__WhereName__:
                self.data[where]=[]
                k={
                    "date":today_date_str(),
                    "where":where
                }
                for gender in self.__Gender__:
                    k[gender]=0
                self.data[where].append(k)
            self.check_date()
            self.save_data()
            
    def save_data(self):#데이터 저장
        with open(self.FilePath,'w') as fk:
            json.dump(self.data, fk,indent=4)
    
    def load_settings(self):
        if os.path.isfile(self.SettingPath):
            with open(self.SettingPath,'r') as stpr:
                self.settings=json.load(stpr)
            
    def check_setting(self,FilePath,HistoryPath,whereList,genderList,imagelist,savetime,savedir,iconpath):
        '''
        세팅 파일에서 읽어온 데이터 체크
        ->만약 내용이 없다면 1번(새로 만들고 저장 후 자기함수호출)
        ->내용이 불완전하다 2번(세팅 읽어온거 다 지우고 자기함수호출)
        ->내용이 맞다면 3번(세이브 파일 요소로 등록 후 함수 나옴)
        '''
        if self.settings=={}:
            self.settings['FilePath']=[]
            self.settings['FilePath'].append({
                'path':FilePath,
                'type':'Data'
            })
            self.settings['FilePath'].append({
                'path':HistoryPath,
                'type':'History'
            })
            self.settings['NameList']=[]
            self.settings['NameList'].append({
                'list':whereList,
                'type':'Place'
            })
            self.settings['NameList'].append({
                'list':genderList,
                'type':'Gender'
            })
            self.settings['ImagePath']=imagelist
            self.settings['DailySave']={
                'time':savetime,
                'dir':os.path.abspath(savedir)
            }
            self.settings['IconPath']=iconpath
            self.save_settings()
            self.check_setting(FilePath,HistoryPath,whereList,genderList,imagelist,savetime,savedir,iconpath)
        else:
            if sorted(list(self.settings.keys()))!=sorted(['FilePath','NameList','ImagePath','DailySave','IconPath']):
                self.settings={}
                self.check_setting(FilePath,HistoryPath,whereList,genderList,imagelist,savetime,savedir,iconpath)
            else:
                self.settings_to_self()
    
    def settings_to_self(self):#세팅 dict로부터 요소들 읽어오기
        for d in self.settings['FilePath']:
            if d['type']=='Data':self.FilePath=d['path']
            elif d['type']=='History':self.HistoryPath=d['path']
        for f in self.settings['NameList']:
            if f['type']=='Place': self.__WhereName__=f['list']
            elif f['type']=='Gender':self.__Gender__=f['list']
        self.imagelist=self.settings['ImagePath']
        self.savetime=self.settings['DailySave']['time']
        self.savedir=self.settings['DailySave']['dir']
        self.icon_path=self.settings['IconPath']
     
    def change_settings(self,**kwargs):#요소를 바꾸고 세팅 dict에 넣은 후 저장, 바뀐 것이 있가면 True 리턴
        ischanged=False
        for k,v in kwargs.items():
            print (k)
            print(v)
            print('===========================')
            k=k.lower()
            if k=='savedir' and os.path.isdir(v):
                print('change savedir')
                self.savedir=v
                self.settings['DailySave']['dir']=self.savedir
                self.save_settings()
                ischanged=True
            if k=='iconpath' and os.path.isfile(v):
                print('change icon path')
                self.icon_path=v
                self.settings['IconPath']=self.icon_path
                self.save_settings()
                ischanged=True
            if k=='savetime' and v is str:
                try:
                    time.strftime(v,'%H:%M')
                    print('change savetime')
                    self.savetime=v
                    self.settings['DailySave']['time']=self.savetime
                    self.save_settings()
                    ischanged=True
                except ValueError:
                    pass
            if k=='imagelist' and type(v)==list:
                for vd in v:#이미지 리스트 체크
                    if os.path.isfile(vd):pass
                    else:v.remove(vd)
                print('change imagelist')
                self.imagelist=v
                self.settings['ImagePath']=self.imagelist
                self.save_settings()
                ischanged=True
            if k=='wherelist' and type(v)==list:
                for vd in v:
                    if type(vd) is not str:
                        v.remove(vd)
                print('change places name')
                self.__WhereName__=v
                for f in self.settings['NameList']:
                    if f['type']=='Place': f['list']=self.__WhereName__
                self.save_settings()
                ischanged=True
            if k=='genderlist' and type(v)==list:
                for vd in v:
                    if type(vd) is not str:
                        v.remove(vd)
                print('change Gender name')
                self.__Gender__=v
                for f in self.settings['NameList']:
                    if f['type']=='Gender': f['list']=self.__Gender__
                self.save_settings()
                ischanged=True
            if k=='filepath':
                print('Change FilePath')
                self.FilePath=v
                for d in self.settings['FilePath']:
                    if d['type']=='Data':d['path']=self.FilePath
                self.save_settings()
                ischanged=True
            if k=='historypath':
                print('Change HistoryPath')
                self.HistoryPath=v
                for d in self.settings['FilePath']:
                    if d['type']=='History':d['path']=self.HistoryPath
                self.save_settings()
                ischanged=True
        return ischanged 
    
    def show_settings(self,*args):
        tmp={}
        for arg in args:
            k=None
            if arg.lower()=='filepath':k=self.FilePath
            elif arg.lower()=='historypath':k=self.HistoryPath
            elif arg.lower()=='wherelist':k=self.__WhereName__
            elif arg.lower()=='genderlist':k=self.__Gender__
            elif arg.lower()=='imagelist':k=self.imagelist
            elif arg.lower()=='savetime':k=self.savetime
            elif arg.lower()=='savedir':k=self.savedir
            elif arg.lower()=='iconpath':k=self.icon_path
            if k!=None:tmp[arg]=k
        return tmp
    def save_settings(self):
        with open(self.SettingPath,'w') as kosk:
            json.dump(self.settings,kosk, indent=4)
    
    def save_history(self):#내역 저장
        with open(self.HistoryPath,'w') as hk:
            json.dump(self.history,hk,indent=4)

    def add_history_member(self,type='push',count=1,gender="",place="",showdates=(today_date_str(),today_date_str())):
        tmp={}
        if type =='built'or type=='init':
            tmp['user']=gp.getuser()
            tmp['time']=now_time_str()
            tmp['type']=type
            tmp['isSucess']=TRUE
        elif type=='push' and gender in self.__Gender__ and place in self.__WhereName__:
            tmp['user']=gp.getuser()
            tmp['time']=now_time_str()
            tmp['type']=type
            tmp['isSucess']=TRUE
            tmp['gender']=gender
            tmp['place']=place
            tmp['count']=count
        elif type=='show' and len(date_range(showdates[0],showdates[1]))>0:
            tmp['user']=gp.getuser()
            tmp['time']=now_time_str()
            tmp['type']=type
            tmp['isSucess']=TRUE
            tmp['dates']=date_range(showdates[0],showdates[1])
        else:
            tmp['user']=gp.getuser()
            tmp['time']=now_time_str()
            tmp['type']=type
            tmp['isSucess']=FALSE
            tmp['count']=count
            if gender!="":tmp['gender']=gender
            if place!="":tmp['place']=place
            self.history['fail'].append(tmp)
            self.save_history()
            return FALSE
        self.history[tmp['type']].append(tmp)
        self.save_history()
        return TRUE

    def check_date(self,date_w=today_date_str()):#오늘 날짜 멤버 체크 및 생성
        today=date_w
        print('check_date')
        print(today)
        self.__Today_Data__=[]
        for where in self.__WhereName__:
            for item in self.data[where]:
                if today==item['date']: 
                    self.__Today_Data__.append(item)
                    break
        print('sada')
        print(self.__Today_Data__)
        if len(self.__Today_Data__)<len(self.__WhereName__):
            hasMember=[]
            for item in self.__Today_Data__:
                hasMember.append(item['where'])
            tmp=list(set(self.__WhereName__)-set(hasMember))
            print('fdasfasd')
            print(tmp)
            for place in tmp:
                kal={
                    "date":today,
                    "where":place
                }
                for gender in self.__Gender__:
                    kal[gender]=0
                self.data[place].append(kal)
            self.check_date(date_w=today)
        elif len(self.__WhereName__)==len(self.__Today_Data__):
            self.save_data()
            return
        else: self.check_date(date_w=today)

    def show_data(self,start=today_date_str(),end=today_date_str(),mode=FALSE):#특정 기간의 데이터 검색
        end_date=datetime.datetime.strptime(end, '%Y-%m-%d')
        start_date=datetime.datetime.strptime(start, '%Y-%m-%d')
        result=[]
        delta=end_date - start_date
        if delta.days==0:
            for place in self.__WhereName__:
                for data in self.data[place]:
                    if datetime.datetime.strptime(data['date'],'%Y-%m-%d')==start_date:
                        result.append(data)
                        break
        elif delta.days>=1:
            for time in [start_date+timedelta(days=i) for i in range(delta.days+1)]:
                for place in self.__WhereName__:
                    for data in self.data[place]:
                        if datetime.datetime.strptime(data['date'],'%Y-%m-%d')==time:
                            result.append(data)
                            break
        if mode:self.add_history_member(type='show',showdates=(start,end))
        return result
                
    def push_data(self,date=today_date_str(),gender="",where="",count=1):#날짜/성별/장소 를 지정 count만큼 추가
        sucess=FALSE
        print('hey')
        if gender not in self.__Gender__:pass#Invalid Gender
        elif where not in self.__WhereName__:pass#Invalid Place
        elif [data for data in self.show_data(date,date) if data['where']==where]==[]:
            print(self.show_data(date,date))#No Member0
            print(where)
        else:
            target=[data for data in self.show_data(date,date) if data['where']==where][0]
            print('kola')
            print(target)
            for member in self.data[where]:
                if member==target:
                    member[gender]+=count
                    if member[gender]<0:
                        member[gender]=0
                    self.data['cumulative']+=count
                    if self.data['cumulative']<=0:
                        self.data['cumulative']=0
                    sucess=TRUE
                    break
        self.save_data()
        self.add_history_member(type='push',gender=gender,place=where,count=count)
        return sucess

    def modify_history_excel(self,excel_path=history_excel_path):
        """
        excel_path : if mode is TRUE excel file will save at excel_path. Defalt :'history_excel.xlsx'
        """
        columns_h={}
        for key in list(self.history.keys()):
            try:
                columns_h[key]=list(self.history[key][0].keys())
            except IndexError:
                columns_h[key]=[]
                
        wb=openpyxl.Workbook()
        for key in list(self.history.keys()):
            ws=wb.create_sheet(key)
            if len(columns_h[key])!=0:
                for cl in range(len(columns_h[key])):
                    ws.cell(row=1,column=cl+2).value=columns_h[key][cl]
                    ws.cell(row=1,column=cl+2).alignment=Alignment(horizontal='center',vertical='center')
                    ws.cell(row=1,column=cl+2).fill=PatternFill(patternType='solid',fgColor=Color('FFC000'))
                for index in range(len(self.history[key])):
                    ws.cell(row=index+2,column=1).value=index+1
                    ws.cell(row=index+2,column=1).alignment=Alignment(horizontal='center',vertical='center')
                    ws.cell(row=index+2,column=1).fill=PatternFill(patternType='solid',fgColor=Color('008000'))
                    ws.cell(row=index+2,column=1).font=Font(bold=TRUE,color=Color('ffffff'))
                    for cl in range(len(columns_h[key])):
                        if type(self.history[key][index][columns_h[key][cl]])!=list:ws.cell(row=index+2,column=cl+2).value=self.history[key][index][columns_h[key][cl]]
                        else:ws.cell(row=index+2,column=cl+2).value=str(self.history[key][index][columns_h[key][cl]])
                        ws.cell(row=index+2,column=cl+2).alignment=Alignment(horizontal='center',vertical='center')
                self.adjust_cell_width(ws)
            else:
                pass
        wb.remove(wb['Sheet'])
        try:
            wb.save(excel_path)
            os.startfile(excel_path)
        except PermissionError:
            messagebox.showerror("PermissionError","파일을 닫아주세요.")

    def adjust_cell_width(self, ws,value=1.2,cl=['A'],ri=[1]):
        """
        Adjust cell width/make (.1)&(A.) thick
        ws : worksheet handle from openpyxl
        cl : options to make other columns thick default : ['a']
        ri : options to make other rows thick default : [1]
        """
        thick_border=Border(left=Side(style='thick'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))
        for col in ws.columns:
            max_length =0
            column_col=col[0].column_letter
            for cell in col:
                cell.alignment=Alignment(horizontal='center',vertical='center')
                try:
                    if cell.column_letter in cl or cell.row in ri:cell.border=thick_border
                    if cell.value==0:cell.value='-'
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    print('kr')
                    pass
            adjusted_width=(max_length+2)*value
            print(column_col)
            print(int(adjusted_width))
            ws.column_dimensions[str(column_col)].width=int(adjusted_width)
        
    def give_savedir(self):#만약 프로그램을 다른 컴퓨터에서 실행한다면 데스크톱 유저 디렉토리도 바뀌므로 그에 대비해서 만든 함수
        tmp=''
        if os.path.isdir(self.savedir):
            pass
        elif self.savedir.split('\\')[-1]=='Desktop':
            tmp=winshell.desktop()
        else:
            tmp='data'
        if tmp!='':self.change_settings(savedir=tmp)
        return self.savedir
    
    def modify_data_excelike(self,selected_dates=[],mode=TRUE,excel_path=excel_path):
        """
        mode : make Excel file default : True
        selected_dates : list of dates that required for showing info DeFault : show_everything
        excel_path : if mode is TRUE excel file will save at excel_path. Defalt :'data_excel.xlsx'
        """
        datesd=[]
        for place in self.__WhereName__:
            for data in self.data[place]:
                if data['date'] not in datesd:datesd.append(data['date'])
        dates=[]
        if selected_dates!=[]:
            for date in datesd:
                if date in selected_dates:
                    dates.append(date)
        else:dates=datesd
        DATE=sorted(dates)
        dates=DATE
        result={}
        columns_d=["Date"]
        columns_name=[]
        for where in self.__WhereName__:
            for gender in self.__Gender__:
                columns_name.append(where+" "+gender)
        columns_d+=columns_name
        columns_d+=['총합']
        for date in range(len(dates)):
            sd=self.show_data(start=dates[date],end=dates[date],mode=TRUE)
            tmp=[dates[date]]
            for place in self.__WhereName__:
                for gender in self.__Gender__:
                    tmp.append(0)
            tmp.append(0)
            for member in sd:
                for i in range(len(self.__WhereName__)):
                    if member['where']==self.__WhereName__[i]:
                        for j in range(len(self.__Gender__)):
                            tmp[i*2+j+1]+=member[self.__Gender__[j]]
                            tmp[-1]+=tmp[i*2+j+1]
            result[str(date+1)]=tmp
        pprint.pprint(result)
        if mode:
            df=pd.DataFrame.from_dict(result,orient="index",columns=columns_d)
            try:
                df.to_excel(excel_path)
                wb=openpyxl.load_workbook(excel_path,data_only=True)
                ws=wb['Sheet1']
                sum_row=[]
                for c in ws.columns:
                    column_ind=c[0].column_letter
                    c[0].alignment=Alignment(horizontal='center',vertical='center')
                    c[0].fill=PatternFill(patternType='solid',fgColor=Color('FFC000'))
                    if column_ind !='A' and column_ind !='B':
                        tmp_add=0
                        for cell in c:
                            cell.alignment=Alignment(horizontal='center',vertical='center')
                            if cell.row!=1:
                                tmp_add+=int(cell.value)
                        sum_row.append(tmp_add)
                    elif column_ind=='A':
                        F=Font(bold=TRUE,color=Color('ffffff'))
                        for cell in c:
                            cell.alignment=Alignment(horizontal='center',vertical='center')
                            cell.fill=PatternFill(patternType='solid',fgColor=Color('008000'))
                            cell.font=F
                    else:
                        F=Font(bold=TRUE,color=Color('ffffff'))
                        for cell in c:
                            cell.alignment=Alignment(horizontal='center',vertical='center')
                            cell.fill=PatternFill(patternType='solid',fgColor=Color('008000'))
                            cell.font=F
                ws.append(["총합"," "]+sum_row)
                kolase=[]
                for kas in range(len(self.__Gender__)):
                    j=0
                    for suro in range(len(sum_row)):
                        if suro%len(self.__Gender__)==kas:
                            j+=sum_row[suro]
                    kolase.append(j)
                kolase[0]-=sum_row[-1]
                kartli=[]
                for genko in range(len(self.__Gender__)):
                    kartli.append(self.__Gender__[genko]+ " 총합 :")
                    kartli.append(kolase[genko])
                place_total=[]
                GT=len(self.__Gender__)
                for plag in range(len(self.__WhereName__)):
                    loki=0
                    for geni in range(GT):
                        loki+=sum_row[plag*GT+geni]
                    place_total.append(self.__WhereName__[plag]+" 총합 :")
                    place_total.append(loki)
                ws.append(["",""]+kartli)
                ws.append(["",""]+place_total)
                self.adjust_cell_width(ws,value=2.4,cl=['A','B'])
                ws.column_dimensions['A'].width=5
                wb.save(excel_path)
                os.startfile(excel_path)
            except PermissionError:
                messagebox.showerror("PermissionError","파일을 닫아주세요.")
        return result

    def columns_name(self):
        columns_name=[]
        for where in self.__WhereName__:
            for gender in self.__Gender__:
                columns_name.append(where+" "+gender)
        return columns_name
    
    def button_event(self,event,where="",gender="",mode=TRUE):
        if where not in self.__WhereName__ or gender not in self.__Gender__:pass
        else:
            if mode==TRUE:self.push_data(gender=gender,where=where)
            else:self.push_data(gender=gender,where=where,count=-1)

    def check_save(self,path=""):
        isValid=FALSE
        if path=="":path=self.FilePath
        if os.path.isfile(path):
            check_tmp={}
            with open(path,'r') as check:
                check_tmp=json.load(check)
            if sorted(list(check_tmp.keys()))==sorted(self.__WhereName__+['cumulative','typecode']) and sorted(list(check_tmp[self.__WhereName__[0]][0].keys()))==sorted(list(self.__Gender__+['date','where'])):
                isValid=TRUE
                print("세이브 파일 무결성 확인")
        else:
            pass
        return isValid

    def check_his(self,path=""):
        isValid=FALSE
        if path=="":path=self.HistoryPath
        if os.path.isfile(path):
            check_tmp={}
            with open(path,'r') as check:
                check_tmp=json.load(check)
            if sorted(list(check_tmp.keys()))==sorted(['init','built','show','push','fail']):
                isValid=TRUE
                print("기록 파일 무결성 확인")
        else:
            pass
        return isValid

    def quit(self,mode=TRUE):
        global FILE_CLOSE
        FILE_CLOSE=mode
        print("자동저장 후 종료")
        self.save_data()
        self.save_history()
        self.save_settings()
        print('저장 완료')
        del self


class MyApp(QWidget):
    def __init__(self,save_file=None,master=None):
        super().__init__()
        self.savefile=save_file
        self.master=master
        self.InitUi()
        self.SetUi()
    
    def InitUi(self):
        self.set_menu()
        self.set_statusbar()
        self.set_widgets()
    
    def set_menu(self):
        pass
    
    def set_statusbar(self):
        pass
    
    def set_widgets(self):
        self.table=QTableWidget(self)
        self.table.resize(1000,250)
        self.setTableWidgetData()
        
    def setTableWidgetData(self):
        dta=self.savefile.modify_data_excelike(selected_dates=[today_date_str()],mode=False)
        df=pandas.DataFrame.from_dict(dta,orient='index')

        self.table.resizeColumnsToContents()
        self.table.resizeRowsToContents()
        self.table.setColumnCount(len(df.columns))
        self.table.setRowCount(len(df.index))
        for i in range(len(df.index)):
            for j in range(len(df.columns)):
                oak=QTableWidgetItem(str(df.iloc[i, j]))
                oak.setTextAlignment(QtCore.Qt.AlignCenter)
                self.table.setItem(i,j,oak)
        self.table.setStyleSheet("border-radius:14px;"
                                 "border-width:2px")
        self.table.setAutoScroll(True)
    def SetUi(self):
        self.setWindowTitle(today_date_str()+'일 기록')
        self.resize(1000,600)
        self.center()
        self.setWindowIcon(QIcon(self.savefile.icon_path))
        self.show()
        
    def center(self):#make window move to center
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())
        
    def closeEvent(self,event):
        event.ignore()
        print('Quit all')
        self.savefile.quit()
        event.accept()

if __name__ == "__main__":
    app=QApplication(sys.argv)
    print(app)
    jd=Json_Data()
    Mw=MyApp(save_file=jd,master=app)
    sys.exit(app.exec_())